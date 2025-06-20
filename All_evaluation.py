import os.path
import pandas as pd
from openpyxl import load_workbook
import glob
import numpy as np

from tqdm.contrib import tzip
# 自作モジュール
from evaluation.PESQ import pesq_evaluation
from evaluation.STOI import stoi_evaluation
from evaluation.SI_SDR import sisdr_evaluation
from mymodule import my_func, const

def split_data(input_data:list, channel:int=0)->list:
    """
    引数で受け取ったtensor型の配列の形状を変形させる[1,音声長×チャンネル数]->[チャンネル数,音声長]

    Parameters
    ----------
    input_data(list[int]):分割する前のデータ[1, 音声長*チャンネル数]
    channels(int):チャンネル数(分割する数)

    Returns
    -------
    split_data(list[float]):分割した後のデータ[チャンネル数, 音声長]
    """
    # print("\nsplit_data")    # 確認用
    """ エラー処理 """
    if channel <= 0:   # channelsの数が0の場合or指定していない場合
        raise ValueError("channels must be greater than 0.")

    # print(f"type(in_tensor):{type(in_tensor)}") # 確認用 # torch.Tensor
    # print(f"in_tensor.shape:{in_tensor.shape}") # 確認用 # [1,音声長×チャンネル数]

    """ 配列の要素数を取得 """
    n = input_data.shape[-1]  # 要素数の取得
    # print(f"n:{n}")         # 確認用 # n=音声長×チャンネル数
    if n % channel != 0:   # 配列の要素数をchannelsで割り切れないとき = チャンネル数が間違っているとき
        raise ValueError("Input array size must be divisible by the number of channels.")

    """ 配列の分割 """
    length = n // channel   # 分割後の1列当たりの要素数を求める
    # print(f"length:{length}")   # 確認用 # 音声長
    trun_input = input_data.T   # 転置
    # print_name_type_shape("trun_tensor", trun_tensor)
    split_input = trun_input.reshape(-1, length) # 分割
    # print_name_type_shape("split_tensor", split_input) # 確認用 # [チャンネル数, 音声長]
    # print("split_data\n")    # 確認用
    return split_input

def make_total_csv(condition:dict, original_path="./evaluation/total_score_original.xlsx", out_dir=os.path.join(const.EVALUATION_DIR, "total_file")):
    """　まとめファイルを作する
    動作確認していない
    
    :param condition: 書き込むファイルのpath
    :return: None
    """
    """ ディレクトリの作成 """
    out_name = f"{condition['speech_type']}_{condition['noise']}_{condition['snr']}dB_{condition['reverbe']}sec"   # ファイル名の作成
    out_path = f"{out_dir}/{out_name}.xlsx"   # 出力パスの作成
    my_func.make_dir(out_path)   # ディレクトリの作成
    print(f"out_path:{out_path}")
    
    if not os.path.isfile(out_path):    # ファイルが存在しない場合
        """ コピー元の読み込み """
        wb = load_workbook(original_path)   # コピー元の読み込み
        sheet = wb["Sheet1"]    # シートの指定
        """ 実験条件の書き込み """
        for idx, item in enumerate(condition.values()):
            cell = sheet.cell(row=2, column=1+idx)
            cell.value = item
        """ 保存 """
        wb.save(out_path)   # ワークブックを別名(out_path)に保存
    return out_path
        
def main(target_dir, estimation_dir, out_path, condition, channel=1):
    """　客観評価を行う


    """
    print("target: ", target_dir)
    print("estimation: ", estimation_dir)

    """ 出力ファイルの作成"""
    my_func.make_dir(out_path)
    with open(out_path, "w") as csv_file:
        csv_file.write(f"target_dir,{target_dir}\nestimation_dir,{estimation_dir}\n")
        csv_file.write(f"{out_path}\ntarget_name,estimation_name,pesq,stoi,sisdr\n")
        
    total_file = make_total_csv(condition=condition)
    

    """ ファイルリストの作成 """
    target_list = my_func.get_file_list(dir_path=target_dir, ext=".wav")
    estimation_list = my_func.get_file_list(dir_path=estimation_dir, ext=".wav")
    # print("target: ",len(target_list))
    # print("estimation: ",len(estimation_list))

    """ 初期化 """
    pesq_sum = 0
    stoi_sum = 0
    sisdr_sum = 0

    for target_file, estimation_file in tzip(target_list, estimation_list):
        """ ファイル名の取得 """
        # print("\n")
        # print("target: ",target_file)
        # print("estimation: ",estimation_file)
        target_name, _ = my_func.get_file_name(target_file)
        estimation_name, _ = my_func.get_file_name(estimation_file)
        """ 音源の読み込み """
        # print(f"target_file:{target_file}")
        target_data, _ = my_func.load_wav(target_file)
        estimation_data, _ = my_func.load_wav(estimation_file)
        # target_data = split_data(target_data, channel=4)
        # estimation_data = split_data(estimation_data, channel=4)
        # print(f"target_data.shape:{target_data.shape}")
        # print(f"estimation_data:{estimation_data.shape}")
        if channel != 1:
            target_data = split_data(target_data, channel)[0]   # 0番目のマイクの音を取得 [音声長 * マイク数] → [音声長]
            # estimation_data = split_data(estimation_data, channel)[0]

        max_length = max(len(target_data), len(estimation_data))
        target_data = np.pad(target_data, [0, max_length - len(target_data)], "constant")
        estimation_data = np.pad(estimation_data, [0, max_length - len(estimation_data)], "constant")
        # min_length = min(len(target_data), len(estimation_data))
        # target_data = target_data[:min_length]
        # estimation_data = estimation_data[:min_length]
        #
        # if len(target_data)>len(estimation_data):
        #     target_data = target_data[:len(estimation_data)]
        # else:
        #     estimation_data = estimation_data[:len(target_data)]

        """ 客観評価の計算 """
        # print()
        pesq_score = pesq_evaluation(target_data, estimation_data)
        # print("pesq: ",pesq_score)
        stoi_score = stoi_evaluation(target_data, estimation_data)
        # print("stoi: ",stoi_score)
        sisdr_score = sisdr_evaluation(target_data, estimation_data)
        # print("sisdr: ",sisdr_score)
        pesq_sum += pesq_score
        stoi_sum += stoi_score
        sisdr_sum += sisdr_score

        """ 出力(ファイルへの書き込み) """
        with open(out_path, "a") as csv_file:  # ファイルオープン
            text = f"{target_name},{estimation_name},{pesq_score},{stoi_score},{sisdr_score}\n"  # 書き込む内容の作成
            csv_file.write(text)  # 書き込み
        # print("save...")

    """ 平均の算出(ファイルへの書き込み) """
    pesq_ave=pesq_sum/len(estimation_list)
    stoi_ave=stoi_sum/len(estimation_list)
    sisdr_ave=sisdr_sum/len(estimation_list)
    with open(out_path, "a") as csv_file:  # ファイルオープン
        text = f"average,,{pesq_ave},{stoi_ave},{sisdr_ave}\n"  # 書き込む内容の作成
        csv_file.write(text)  # 書き込み
    """ まとめファイルへの書き込み """
    work_book = load_workbook(total_file)
    sheet = work_book["Sheet1"]
    max_row = my_func.get_max_row(sheet)
    # print(f"max_row:{max_row}")
    sheet.cell(row=max_row, column=1).value = out_path
    sheet.cell(row=max_row, column=2).value = float(pesq_ave)
    sheet.cell(row=max_row, column=3).value = float(stoi_ave)
    sheet.cell(row=max_row, column=4).value = float(sisdr_ave)
    sheet.cell(row=max_row, column=5).value = target_dir
    sheet.cell(row=max_row, column=6).value = estimation_dir
    work_book.save(total_file)

    print("")
    print(f"PESQ : {pesq_ave}")
    print(f"STOI : {stoi_ave}")
    print(f"SI-SDR : {sisdr_ave}")
    # print("pesq end")

if __name__ == "__main__":
    print("start evaluation")
    condition = {"speech_type": "subset_DEMAND",
                 "noise": "hoth",
                 "snr": 10,
                 "reverbe": 5}

    # model_list = ["A", "C", "D", "E"]
    # wave_type_list = ["noise_only", "reverbe_only", "noise_reverbe"]    # "noise_only", "reverbe_only", "noise_reverbe"
    # # for reverbe in range(1, 6):
    # for wave_type in wave_type_list:
    #     for model_type in model_list:
    #         # a = f"{const.OUTPUT_WAV_DIR}/subset_DEMAND_hoth_1010dB_05sec_4ch_10cm/{model_type}/Right/{wave_type}"
    #         # C:\Users\kataoka-lab\Desktop\sound_data\mix_data\subset_DEMAND_hoth_1010dB_05sec_4ch_10cm\Right\test\clean
    #         main(target_dir=f"{const.MIX_DATA_DIR}\\subset_DEMAND_hoth_1010dB_05sec_4ch_10cm\\Right\\test\\clean",
    #              estimation_dir=f"{const.OUTPUT_WAV_DIR}/subset_DEMAND_hoth_1010dB_05sec_4ch_10cm/{model_type}/Right/{wave_type}",
    #              out_path=f"{const.EVALUATION_DIR}\\{model_type}_{wave_type}.csv",
    #              condition=condition,
    #              channel=4)
    # C:\Users\kataoka-lab\Desktop\sound_data\mix_data\subset_DEMAND_1ch\condition_1\test\noise_reverbe
    i = 1
    # {const.OUTPUT_WAV_DIR}/URelNet/subset_DEMAND_hoth_1010dB_05sec_1ch_0cm/noise_reverbe
    main(target_dir=f"{const.MIX_DATA_DIR}/subset_DEMAND_hoth_1010dB_05sec_1ch_0cm/test/clean",
         estimation_dir=f"{const.OUTPUT_WAV_DIR}/URelNet/subset_DEMAND_hoth_1010dB_05sec_1ch_0cm/reverbe_only",
         out_path=f"{const.EVALUATION_DIR}/URelNet/subset_DEMAND_hoth_1010dB_05sec_1ch_0cm/reverbe_only.csv",
         condition=condition,
         channel=1)
